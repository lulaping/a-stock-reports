"""生成超短复盘Excel模板 — 清单1~7"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 样式定义
header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='E94560', end_color='E94560', fill_type='solid')
sub_font = Font(name='Microsoft YaHei', size=11, bold=True)
sub_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
cell_font = Font(name='Microsoft YaHei', size=10)
red_font = Font(name='Microsoft YaHei', size=10, color='CF222E', bold=True)
green_font = Font(name='Microsoft YaHei', size=10, color='1A7F37')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_range(ws, row_start, row_end, col_start, col_end, font=None, fill=None, alignment=None):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if font: cell.font = font
            if fill: cell.fill = fill
            if alignment: cell.alignment = alignment
            cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ====== Sheet 1: 清单1 每日复盘 ======
ws1 = wb.active
ws1.title = "清单1-每日复盘"
set_col_widths(ws1, [18, 14, 20, 14, 18, 20])

# 标题
ws1.merge_cells('A1:F1')
ws1['A1'] = '清单1：每日复盘（15:00收盘后）'
ws1['A1'].font = header_font
ws1['A1'].fill = header_fill
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 30

# 日期行
ws1.merge_cells('A2:F2')
ws1['A2'] = '复盘日期：__________    星期___    交易员：__________'
ws1['A2'].font = sub_font
ws1['A2'].fill = sub_fill
ws1['A2'].alignment = left_wrap

# 情绪仪表盘
headers = ['指标', '今日数值', '健康区间', '判定', '备注', '昨日对比']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws1.cell(row=3, column=i).alignment = center
    ws1.cell(row=3, column=i).border = thin_border

data = [
    ['涨停家数', '____', '≥60正常；<40冰点', '□好 □差', '', ''],
    ['跌停家数', '____', '≤5正常；>10退潮', '□好 □差', '', ''],
    ['炸板率', '____%', '<25%好；>35%差', '□好 □差', '', ''],
    ['1进2晋级率', '____/____', '>20%正常', '____%', '', ''],
    ['2进3晋级率', '____/____', '>30%正常', '____%', '', ''],
    ['3进4+晋级率', '____/____', '统计≥3板存活', '____%', '', ''],
    ['最高连板', '____板', '≥4板情绪好', '□好 □差', '', ''],
    ['成交额', '____万亿', '≥2万亿增量', '□增量 □缩量', '', ''],
    ['上证指数', '____', '涨跌：____%', '□涨 □跌', '', ''],
    ['创业板指', '____', '涨跌：____%', '□涨 □跌', '', ''],
]

for i, row_data in enumerate(data):
    for j, val in enumerate(row_data):
        cell = ws1.cell(row=4 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = center if j < 4 else left_wrap
        cell.border = thin_border

# 定性判断
qual_row = 4 + len(data) + 1
ws1.merge_cells(f'A{qual_row}:F{qual_row}')
ws1.cell(row=qual_row, column=1, value='定性判断').font = sub_font
ws1.cell(row=qual_row, column=1).fill = sub_fill
ws1.cell(row=qual_row, column=1).alignment = left_wrap

qual_items = [
    '□趋势日 □连板日    明日策略：___________（趋势日→低吸核心；连板日→打板接力）',
    '最强主线板块：___________（板块内≥3只涨停才算成建制）',
    '今日弱转强标的：___________（昨日烂板/分歧，今日拉红封板）',
    '今日异动公告：____只  → 若≥3只，次日退潮概率高，降仓',
    '明日情绪温度预估：____°C  → 参考仓位：<50°C≤3成 / 50-90°C 5成 / 90-120°C 7成 / >120°C 5成',
    '今日赚钱效应：___________（板块/个股/模式）',
    '今日亏钱效应：___________（板块/个股/模式）',
]
for i, item in enumerate(qual_items):
    ws1.merge_cells(f'A{qual_row + 1 + i}:F{qual_row + 1 + i}')
    ws1.cell(row=qual_row + 1 + i, column=1, value=item).font = cell_font
    ws1.cell(row=qual_row + 1 + i, column=1).alignment = left_wrap


# ====== Sheet 2: 清单2 晚间选股 ======
ws2 = wb.create_sheet("清单2-晚间选股")
set_col_widths(ws2, [8, 16, 14, 14, 14, 14, 16, 14, 22])

ws2.merge_cells('A1:I1')
ws2['A1'] = '清单2：晚间选股（20:00 — 筛选次日候选）'
ws2['A1'].font = header_font
ws2['A1'].fill = header_fill
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 30

# 淘汰检查
ws2.merge_cells('A2:I2')
ws2['A2'] = '淘汰检查：□大股东质押>50%  □商誉/净资产>50%  □已发澄清公告  □龙虎榜一日游  □同板块高标断板'
ws2['A2'].font = Font(name='Microsoft YaHei', size=10, color='CF222E')

headers2 = ['级别', '股票代码', '股票名称', '连板数', '涨停原因', '板块', '是否弱转强', '竞价量比', '选股逻辑/备注']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h).font = Font(name='Microsoft YaHei', size=10, bold=True)
    ws2.cell(row=3, column=i).fill = PatternFill(start_color='E8E8E8', fill_type='solid')
    ws2.cell(row=3, column=i).alignment = center
    ws2.cell(row=3, column=i).border = thin_border

# S/A/B 分级填充行
for level, color, start_row in [('S级', '1A7F37', 4), ('A级', '9A6700', 10), ('B级', 'CF222E', 16)]:
    for n in range(5):
        r = start_row + n
        ws2.cell(row=r, column=1, value=level if n == 0 else '').font = Font(name='Microsoft YaHei', size=10, bold=True, color=color)
        ws2.cell(row=r, column=1).alignment = center
        for c in range(2, 10):
            ws2.cell(row=r, column=c).font = cell_font
            ws2.cell(row=r, column=c).alignment = center
            ws2.cell(row=r, column=c).border = thin_border

# ====== Sheet 3: 清单3 竞价阶段 ======
ws3 = wb.create_sheet("清单3-竞价阶段")
set_col_widths(ws3, [16, 28, 28, 14, 14])

ws3.merge_cells('A1:E1')
ws3['A1'] = '清单3：竞价阶段（9:15-9:25）'
ws3['A1'].font = header_font
ws3['A1'].fill = header_fill
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 30

headers3 = ['检查项', '通过标准', '不通过则', '结果', '备注']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws3.cell(row=2, column=i).fill = sub_fill
    ws3.cell(row=2, column=i).alignment = center
    ws3.cell(row=2, column=i).border = thin_border

checklist3 = [
    ['高开幅度', '一进二2-5%；高标1-3%', '＞7%高开→放弃；低开＞3%→放弃', '', ''],
    ['竞价量比', '竞价成交/昨日成交 5-15%', '＜3%→放弃；＞25%→观望', '', ''],
    ['竞价形态', '低开→9:25拉高(分歧转一致)', '高开→9:20后持续撤单→放弃', '', ''],
    ['20cm票', '9:20-9:25封单递增', '大幅低开→放弃', '', ''],
    ['板块确认', '同板块≥2只竞价高开', '独苗高开→降半仓', '', ''],
    ['弱转强信号', '昨日烂板+今日竞价低开拉红', '5分钟未拉红→放弃', '', ''],
    ['大盘状态', '竞价高开且未翻绿', '大盘竞价低开翻绿→推迟30分钟', '', ''],
]
for i, row_data in enumerate(checklist3):
    for j, val in enumerate(row_data):
        cell = ws3.cell(row=3 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = left_wrap
        cell.border = thin_border
    ws3.row_dimensions[3 + i].height = 28

# ====== Sheet 4: 清单4 盘中执行 ======
ws4 = wb.create_sheet("清单4-盘中执行")
set_col_widths(ws4, [16, 20, 30, 14, 14, 14])

ws4.merge_cells('A1:F1')
ws4['A1'] = '清单4：盘中执行（9:30-15:00）'
ws4['A1'].font = header_font
ws4['A1'].fill = header_fill
ws4['A1'].alignment = center
ws4.row_dimensions[1].height = 30

headers4 = ['时间窗口', '动作', '纪律', '持仓1', '持仓2', '持仓3']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=2, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws4.cell(row=2, column=i).fill = sub_fill
    ws4.cell(row=2, column=i).alignment = center
    ws4.cell(row=2, column=i).border = thin_border

checklist4 = [
    ['9:30-9:35', '弱转强拉红确认', '竞价低开5分钟内拉红→半仓试；未拉红→放弃', '', '', ''],
    ['9:35-10:00', '确认封板质量', '封单/成交＞3x→可加仓；开板≥3次→减仓', '', '', ''],
    ['10:00-10:30', '早盘封板可追', '10:30前封板→质量好；午后封板→只低吸不追', '', '', ''],
    ['10:30-14:30', '持有观察', '只低吸不追高；监控同板块高标', '', '', ''],
    ['14:30后', '尾盘不碰', '14:30后才封板→一律不参与', '', '', ''],
    ['全天', '监控高标断板', '同板块最高标断板→立即减仓', '', '', ''],
    ['全天', '监控异动公告', '持仓股盘中发公告→立即清仓', '', '', ''],
    ['全天', '监控大盘', '大盘跳水翻绿→减仓至3成', '', '', ''],
]
for i, row_data in enumerate(checklist4):
    for j, val in enumerate(row_data):
        cell = ws4.cell(row=3 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = left_wrap
        cell.border = thin_border
    ws4.row_dimensions[3 + i].height = 28

# ====== Sheet 5: 清单5 退出纪律 ======
ws5 = wb.create_sheet("清单5-退出纪律")
set_col_widths(ws5, [6, 40, 22, 14, 14, 14])

ws5.merge_cells('A1:F1')
ws5['A1'] = '清单5：退出纪律（无例外）'
ws5['A1'].font = header_font
ws5['A1'].fill = header_fill
ws5['A1'].alignment = center
ws5.row_dimensions[1].height = 30

headers5 = ['#', '铁律', '触发条件', '今日是否触发', '持仓1', '持仓2']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws5.cell(row=2, column=i).fill = sub_fill
    ws5.cell(row=2, column=i).alignment = center
    ws5.cell(row=2, column=i).border = thin_border

rules5 = [
    ['1', '断板次日必须止损', '昨日涨停→今日未封板', '', '', ''],
    ['2', '一字板不追', '一字板开板→不买入', '', '', ''],
    ['3', '同梯队2只断板→全清', '同梯队≥2只同时断板', '', '', ''],
    ['4', '单票亏损＞-5%→止损', '持仓浮亏超过5%', '', '', ''],
    ['5', '大盘跳水翻绿→减仓至3成', '指数翻绿且持续走弱', '', '', ''],
    ['6', '盘中异动公告→立即清仓', '持仓股盘中发公告', '', '', ''],
    ['7', '14:30后封板→不参与', '尾盘偷袭涨停', '', '', ''],
    ['8', '＞120°C高潮→降仓至5成', '情绪温度超120°C', '', '', ''],
]
for i, row_data in enumerate(rules5):
    for j, val in enumerate(row_data):
        cell = ws5.cell(row=3 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = left_wrap
        cell.border = thin_border
    ws5.row_dimensions[3 + i].height = 24

# ====== Sheet 6: 清单6 仓位管理 ======
ws6 = wb.create_sheet("清单6-仓位管理")
set_col_widths(ws6, [18, 14, 14, 14, 14, 20])

ws6.merge_cells('A1:F1')
ws6['A1'] = '清单6：仓位管理'
ws6['A1'].font = header_font
ws6['A1'].fill = header_fill
ws6['A1'].alignment = center
ws6.row_dimensions[1].height = 30

headers6 = ['情绪温度', '仓位上限', '单票上限', '单日交易上限', '今日实际', '备注']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=2, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws6.cell(row=2, column=i).fill = sub_fill
    ws6.cell(row=2, column=i).alignment = center
    ws6.cell(row=2, column=i).border = thin_border

data6 = [
    ['＜50°C（冰点）', '≤3成', '≤1成', '≤2只', '', ''],
    ['50-90°C（温和）', '≤5成', '≤1.5成', '≤3只', '', ''],
    ['90-120°C（活跃）', '≤7成', '≤2成', '≤3只', '', ''],
    ['＞120°C（高潮）', '≤5成（降仓）', '≤1.5成', '≤2只', '', '预留3成现金'],
]
for i, row_data in enumerate(data6):
    for j, val in enumerate(row_data):
        cell = ws6.cell(row=3 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = center if j < 4 else left_wrap
        cell.border = thin_border
    ws6.row_dimensions[3 + i].height = 24

# ====== Sheet 7: 清单7 三类模式速查 ======
ws7 = wb.create_sheet("清单7-三类模式速查")
set_col_widths(ws7, [16, 30, 28, 26, 14])

ws7.merge_cells('A1:E1')
ws7['A1'] = '清单7：三类高胜率模式快速识别卡'
ws7['A1'].font = header_font
ws7['A1'].fill = header_fill
ws7['A1'].alignment = center
ws7.row_dimensions[1].height = 30

headers7 = ['模式', '识别条件', '买入时机', '退出条件', '今日候选']
for i, h in enumerate(headers7, 1):
    ws7.cell(row=2, column=i, value=h).font = Font(name='Microsoft YaHei', size=11, bold=True)
    ws7.cell(row=2, column=i).fill = sub_fill
    ws7.cell(row=2, column=i).alignment = center
    ws7.cell(row=2, column=i).border = thin_border

data7 = [
    ['① 弱转强反包', '昨日烂板/炸板/分歧\n+今日竞价低开→分时拉红', '分时拉红确认后半仓\n封板加仓', '次日不封板即走\n炸板走', ''],
    ['② 板块成建制', '板块≥3只涨停\n+高标带队+加速期', '竞价确认板块高开\n+个股过竞价清单', '板块高标断板→全清', ''],
    ['③ 产业资本催化', '股权转让/大基金入股\n/定点中标公告', '公告后首板打板\n/次日竞价确认', '公告利好兑现\n(2-3板后)分批止盈', ''],
]
for i, row_data in enumerate(data7):
    for j, val in enumerate(row_data):
        cell = ws7.cell(row=3 + i, column=j + 1, value=val)
        cell.font = cell_font
        cell.alignment = left_wrap
        cell.border = thin_border
    ws7.row_dimensions[3 + i].height = 50

# 保存
output_path = 'e:/Trae/超短复盘模板_清单1-7.xlsx'
wb.save(output_path)
print(f'Excel模板已生成: {output_path}')