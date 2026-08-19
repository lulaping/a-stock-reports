# -*- coding: utf-8 -*-
"""
连板前瞻验证数据库自动更新（L8 · 自动版 v3.0）
==========================================
功能：收盘后自动抓取东方财富涨停/跌停/炸板池 → 计算情绪温度计 → 验证候选池 → 刷新 Excel

自动维护的 Sheet：
  - 行情快照   每日涨停/跌停/炸板/晋级率/最高标/情绪温度/总开关判定（已存在则就地更新）
  - 连板梯队   每日连板股明细（板数/板块/换手/封板时间/炸板次数）（该日数据每日重建）
  - 候选明细   自动验证"候选池配置.json"中的候选（晋级/断板/涨跌幅/可执行收益/断板亏损）
  - 分级命中率 按验证日自动追加 S/A/B 级命中统计（该日已存在则跳过，插到"累计"行前）
  - 方向命中率 按验证日自动追加各方向命中率（该日已存在则跳过）
人工维护的 Sheet（验证复盘报告中填写）：规则验证 / 竞价信号

用法：
  python 更新连板前瞻验证数据库.py              # 自动抓取最近一个交易日并更新（配合 15:20 计划任务）
  python 更新连板前瞻验证数据库.py 20260820     # 指定日期（补录历史）
  python 更新连板前瞻验证数据库.py --show       # 查看配置中待验证的候选
  python 更新连板前瞻验证数据库.py --add 600613 神奇制药 A 5 医药 20260820 化学制药  # 添加候选
  python 更新连板前瞻验证数据库.py --init-task  # 安装 Windows 计划任务（每日 15:20 自动运行）

候选池配置：每天做前瞻时用 --add 把候选写入 候选池配置.json；次日收盘脚本自动验证并刷新 Excel。
注：东财涨停/跌停/炸板池与同花顺口径存在差异（东财涨停家数偏少），温度计口径以脚本抓取值为准。
"""
import json, os, sys, time, datetime
import requests
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(BASE, "连板前瞻验证数据库.xlsx")
CONFIG_JSON = os.path.join(BASE, "候选池配置.json")

UT = "7eea3edcaed734bea9cbfc24409ed989"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
           "Referer": "https://quote.eastmoney.com/"}
POOL_URL = "https://push2ex.eastmoney.com/getTopic{kind}Pool"
QUOTE_URL = "https://push2.eastmoney.com/api/qt/ulist.np/get"

HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
RED = Font(color="CC0000", bold=True)
GREEN = Font(color="008000", bold=True)
THIN = Side(style="thin", color="444444")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SNAP_WIDTHS = [10, 9, 9, 9, 8, 9, 9, 8, 10, 9, 6, 14, 16, 14]
LB_WIDTHS = [10, 9, 10, 8, 12, 9, 9, 9, 12, 8]
CAND_WIDTHS = [10, 10, 8, 8, 10, 10, 8, 12, 8, 14, 10, 10, 9, 24]


def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"


# ================= 数据抓取 =================

def fetch_pool(kind, date):
    """kind: 'zt' 涨停 / 'dt' 跌停 / 'zb' 炸板。返回 (list[dict], 总数)"""
    # 注意：炸板池必须用 sort=fbt:asc，否则返回空 pool；涨停/跌停池用 sort=fund:asc
    sort = "fbt:asc" if kind == "zb" else "fund:asc"
    url = (f"{POOL_URL.format(kind=kind.upper())}?ut={UT}&dpt=wz.ztzt"
           f"&Pageindex=0&pagesize=1000&sort={sort}&date={date}")
    last_err = None
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            j = r.json()
            data = j.get("data") or {}
            pool = data.get("pool") or []
            return pool, data.get("tc", len(pool))
        except Exception as e:
            last_err = e
            time.sleep(2)
    raise RuntimeError(f"抓取{kind}池失败: {last_err}")


def fetch_quotes(secids):
    """批量查询个股行情。secids: ['1.600613','0.300313'...] → dict code->quote"""
    if not secids:
        return {}
    out = {}
    for i in range(0, len(secids), 50):
        batch = secids[i:i + 50]
        for attempt in range(3):
            try:
                r = requests.get(QUOTE_URL, params={
                    "secids": ",".join(batch),
                    "fields": "f2,f3,f8,f12,f14,f15,f16,f17,f18"
                }, headers=HEADERS, timeout=20)
                arr = ((r.json().get("data") or {}).get("diff") or [])
                for q in arr:
                    code = str(q.get("f12", ""))
                    out[code] = {
                        "name": q.get("f14"), "price": (q.get("f2") or 0) / 100,
                        "pct": (q.get("f3") or 0) / 100, "hs": q.get("f8") or 0,
                        "high": (q.get("f15") or 0) / 100, "low": (q.get("f16") or 0) / 100,
                        "open": (q.get("f17") or 0) / 100, "preclose": (q.get("f18") or 0) / 100,
                    }
                break
            except Exception as e:
                print(f"  批量行情查询失败(第{attempt+1}次): {e}")
                time.sleep(2)
        time.sleep(0.3)
    return out


def to_secid(code):
    code = str(code).zfill(6)
    return ("1." if code[0] in "569" else "0.") + code


# ================= 指标计算 =================

def compute_snapshot(zt_pool, dt_pool, zb_pool, prev_lbc_total):
    """计算情绪温度计指标。prev_lbc_total: 昨日连板总数（用于晋级率，无则 None）"""
    zt = len(zt_pool)
    dt = len(dt_pool)
    zb = len(zb_pool)
    max_lbc = max((int(p.get("lbc") or 0) for p in zt_pool), default=0)
    today_lbc = sum(1 for p in zt_pool if int(p.get("lbc") or 0) >= 2)
    zb_rate = round(zb / (zt + zb) * 100, 1) if (zt + zb) else 0.0
    jj_rate = round(today_lbc / prev_lbc_total * 100, 1) if prev_lbc_total else None

    # 温度计打分（L8 规则）
    s_zt = 30 if zt >= 100 else 20 if zt >= 60 else 10 if zt >= 40 else 0
    s_dt = 20 if dt == 0 else 15 if dt <= 5 else 8 if dt <= 15 else 0
    s_jj = (20 if jj_rate is None else
            20 if jj_rate >= 40 else 14 if jj_rate >= 25 else 8 if jj_rate >= 15 else 0)
    s_gd = 15 if max_lbc >= 7 else 10 if max_lbc >= 5 else 6 if max_lbc >= 3 else 2
    s_zb = 15 if zb_rate <= 20 else 10 if zb_rate <= 30 else 5 if zb_rate <= 40 else 0
    temp = s_zt + s_dt + s_jj + s_gd + s_zb

    # 退潮确认：涨停<60 或 高标(≥5板)跌停≥2
    # 注意：跌停池连板字段为 days（涨停池才是 lbc）
    gao_biao_dt = [p["n"] for p in dt_pool if int(p.get("days") or 0) >= 5]
    tui = (zt < 60) or (len(gao_biao_dt) >= 2)

    zone = "高潮" if temp >= 80 else "正常" if temp >= 40 else "冰点"
    verdict = ("空仓观望(退潮确认)" if tui else
               "高潮≤2成" if zone == "高潮" else
               "正常≤5成" if zone == "正常" else "冰点3-5成低吸")
    return {
        "zt": zt, "dt": dt, "zb": zb, "max_lbc": max_lbc, "today_lbc": today_lbc,
        "jj_rate": jj_rate, "zb_rate": zb_rate, "temp": temp, "zone": zone,
        "verdict": verdict, "tui": tui, "gao_biao_dt": gao_biao_dt,
    }


# ================= 候选配置 =================

def load_config():
    if not os.path.exists(CONFIG_JSON):
        save_config({"候选池": {}})
    with open(CONFIG_JSON, encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_JSON, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def add_candidate(code, name, level, prev_boards, direction, date, board=None):
    cfg = load_config()
    cfg.setdefault("候选池", {}).setdefault(date, []).append({
        "代码": code, "名称": name, "级别": level, "昨板数": prev_boards,
        "方向": direction, "板块": board or ""
    })
    save_config(cfg)
    print(f"已添加候选 {date} {code} {name} 级别{level} 昨{prev_boards}板 方向[{direction}] 板块[{board or ''}]")


# ================= Excel 读写 =================

def _style_headers(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
        c = ws.cell(row=1, column=i)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    ws.freeze_panes = "A2"


def _style_body(ws, nrows, ncols, verdict_col=None):
    for r in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col); c.border = BORDER; c.font = Font(size=10)
            v = str(c.value or "")
            if verdict_col and col == verdict_col:
                if "晋级" in v: c.font = GREEN
                elif "断板" in v or "跌停" in v: c.font = RED


def _sort_sheet_by_date(ws, ncols):
    """按第1列日期升序重排数据行并统一样式（保证行序恒定）"""
    data = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            data.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])
    if not data:
        return
    data.sort(key=lambda row: str(row[0] or "").replace("-", ""))
    for r in range(ws.max_row, 1, -1):
        ws.delete_rows(r)
    for row in data:
        ws.append(row)
    _style_body(ws, len(data), ncols)


def get_or_sheet(wb, name):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)


def ensure_snap_headers(ws):
    if ws.max_row < 2 or ws.cell(row=1, column=1).value != "日期":
        ws.delete_rows(1, ws.max_row)
        ws.append(["日期", "涨停家数", "跌停家数", "炸板家数", "炸板率%", "连板家数",
                   "昨日连板", "晋级率%", "最高标板数", "情绪温度", "分区", "总开关判定",
                   "高标跌停", "备注"])
        _style_headers(ws, SNAP_WIDTHS)


def ensure_lb_headers(ws):
    if ws.max_row < 2 or ws.cell(row=1, column=1).value != "日期":
        ws.delete_rows(1, ws.max_row)
        ws.append(["日期", "代码", "名称", "连板数", "板块", "换手率%", "封板时间",
                   "炸板次数", "流通市值(亿)", "状态"])
        _style_headers(ws, LB_WIDTHS)


def ensure_cand_headers(ws):
    first = ws.cell(row=1, column=1).value
    if first is None or str(first).strip() == "":
        ws.delete_rows(1, ws.max_row)
        ws.append(["验证日", "前瞻发布日", "分级", "代码", "名称", "昨日状态", "换手率%",
                   "板块", "方向", "今日结果", "收盘涨跌幅%", "可执行收益%", "断板亏损%", "归因/备注"])
        _style_headers(ws, CAND_WIDTHS)
    elif first == "验证日" and ws.max_column == 13 and ws.cell(row=1, column=9).value == "今日结果":
        # 旧版13列格式 → 迁移到新版14列（第9列补"方向"）
        ws.insert_cols(9)
        ws.cell(row=1, column=9, value="方向")
        _style_headers(ws, CAND_WIDTHS)


# ================= 候选验证 =================

def verify_candidates(cands, zt_pool, dt_pool, quotes, date, prev_date):
    """验证候选：晋级/断板 + 涨跌幅 + 可执行收益 + 断板亏损。返回待写Excel的行列表"""
    zt_codes = {str(p["c"]): p for p in zt_pool}
    dt_codes = {str(p["c"]): p for p in dt_pool}
    rows = []
    for c in cands:
        code = str(c["代码"]).zfill(6)
        name = c.get("名称", "")
        lvl = c.get("级别", "B")
        boards = int(c.get("昨板数") or 1)
        direction = c.get("方向", "")
        board = c.get("板块", "")
        q = quotes.get(code, {})
        pct = q.get("pct")
        preclose = q.get("preclose")
        open_pct = round((q.get("open", 0) / preclose - 1) * 100, 2) if preclose else None
        if code in zt_codes:
            p = zt_codes[code]
            new_lbc = int(p.get("lbc") or boards + 1)
            result = f"✓ 晋级{new_lbc}板"
            hs = round(p.get("hs") or q.get("hs") or 0, 2)
            exec_ret = None; loss = None
            if open_pct is not None and -2 <= open_pct <= 5:
                exec_ret = round((pct or 0) - open_pct, 2)
            note = f"封板{p.get('fbt') or ''} 换手{hs}% 炸板{int(p.get('zbc') or 0)}次"
        elif code in dt_codes:
            result = "✗ 跌停断板"
            hs = round(q.get("hs") or 0, 2)
            exec_ret = None
            loss = pct if pct is not None else -10.0
            note = "跌停"
        else:
            result = "✗ 断板"
            hs = round(q.get("hs") or 0, 2)
            exec_ret = None
            loss = pct if pct is not None else None
            note = ""
        rows.append([date, prev_date, lvl, code, name, f"{boards}板", hs, board, direction,
                     result, pct, exec_ret, loss, note])
    return rows


# ================= 汇总Sheet自动追加 =================

def append_graded_stats(ws, rows, date_str, prev_date, snap_metrics):
    """分级命中率：验证日不存在时追加一行（插到"累计"行之前）"""
    if not rows:
        return
    d = fmt_date(date_str)
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").replace("-", "") == date_str:
            return  # 该日已统计
    cnt = {}; hit = {}
    for row in rows:
        lvl = str(row[2]); res = str(row[9])
        cnt[lvl] = cnt.get(lvl, 0) + 1
        if "✓" in res:
            hit[lvl] = hit.get(lvl, 0) + 1
    def fmt(k):
        return (f"{hit.get(k,0)}/{cnt.get(k,0)} ({round(hit.get(k,0)/cnt[k]*100)}%)"
                if cnt.get(k) else "-")
    core_t = cnt.get("S", 0) + cnt.get("A", 0) + cnt.get("B", 0)
    core_h = hit.get("S", 0) + hit.get("A", 0) + hit.get("B", 0)
    core = f"{core_h}/{core_t} ({round(core_h/core_t*100)}%)" if core_t else "-"
    total_h = sum(hit.values())
    total = f"{total_h}/{len(rows)} ({round(total_h/len(rows)*100)}%)" if rows else "-"
    env = f"{snap_metrics['zone']} {snap_metrics['temp']}分"
    vals = [d, fmt_date(prev_date) if prev_date else "-", env,
            fmt("S"), fmt("A"), fmt("B"), core, total, ""]
    ins_row = ws.max_row + 1
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "") == "累计":
            ins_row = r
            break
    ws.insert_rows(ins_row)
    for i, v in enumerate(vals, 1):
        ws.cell(row=ins_row, column=i, value=v)
    _style_body(ws, ws.max_row - 1, 9)


def append_direction_stats(ws, rows, date_str):
    """方向命中率：验证日不存在时按方向追加多行"""
    if not rows:
        return
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").replace("-", "") == date_str:
            return  # 该日已统计
    by_dir = OrderedDict()
    for row in rows:
        by_dir.setdefault(row[8] or "未分类", []).append(row)
    for d, lst in by_dir.items():
        names = "".join(f"{row[4]}{'✓' if '✓' in str(row[9]) else '✗'}" for row in lst)
        hits = sum(1 for row in lst if "✓" in str(row[9]))
        rate = f"{hits}/{len(lst)} ({round(hits/len(lst)*100)}%)"
        lvls = "/".join(sorted(set(str(row[2]) for row in lst)))
        ws.append([fmt_date(date_str), d, "", names, rate, lvls, ""])
    _style_body(ws, ws.max_row - 1, 7)


# ================= 主流程 =================

def last_trading_day(today):
    """找最近交易日（回退跳过周末；节假日由接口空数据兜底）"""
    d = today
    for _ in range(7):
        if d.weekday() < 5:
            return d
        d -= datetime.timedelta(days=1)
    return d


def prev_session(date_str):
    d = datetime.datetime.strptime(date_str, "%Y%m%d")
    return last_trading_day(d - datetime.timedelta(days=1)).strftime("%Y%m%d")


def main(date_str=None):
    today = datetime.date.today()
    if not date_str:
        date_str = today.strftime("%Y%m%d")
        # 若今天非交易日或盘中，自动回退到上一交易日收盘数据
        if today.weekday() >= 5 or datetime.datetime.now().hour < 15:
            date_str = last_trading_day(today - (datetime.timedelta(days=0 if today.weekday() >= 5 else 1))).strftime("%Y%m%d")
    print(f"== 抓取日期: {date_str} ==")

    # 1. 抓取三池
    zt_pool, zt_tc = fetch_pool("zt", date_str)
    dt_pool, dt_tc = fetch_pool("dt", date_str)
    zb_pool, zb_tc = fetch_pool("zb", date_str)
    if not zt_pool and not dt_pool:
        print(f"  {date_str} 无行情数据（可能非交易日），跳过。")
        return
    print(f"  涨停{len(zt_pool)} 跌停{len(dt_pool)} 炸板{len(zb_pool)}")

    # 2. 加载Excel，取上一快照行的连板总数算晋级率
    prev_total_lbc = None
    prev_date = prev_session(date_str)
    wb = load_workbook(OUT_XLSX) if os.path.exists(OUT_XLSX) else Workbook()
    if wb.sheetnames and wb.sheetnames[0] == "Sheet":
        del wb["Sheet"]
    snap = get_or_sheet(wb, "行情快照")
    ensure_snap_headers(snap)
    for r in range(2, snap.max_row + 1):
        pd_ = str(snap.cell(row=r, column=1).value or "").replace("-", "")
        if pd_ and pd_ < date_str:
            prev_total_lbc = snap.cell(row=r, column=6).value or 0
            break

    # 3. 指标计算 + 更新行情快照（就地更新，避免脏数据残留）
    snap_metrics = compute_snapshot(zt_pool, dt_pool, zb_pool, prev_total_lbc)
    print(f"  温度 {snap_metrics['temp']}/100 {snap_metrics['zone']} → {snap_metrics['verdict']}")
    snap_vals = [fmt_date(date_str), snap_metrics["zt"], snap_metrics["dt"], snap_metrics["zb"],
                 snap_metrics["zb_rate"], snap_metrics["today_lbc"], prev_total_lbc,
                 snap_metrics["jj_rate"], snap_metrics["max_lbc"], snap_metrics["temp"],
                 snap_metrics["zone"], snap_metrics["verdict"],
                 ",".join(snap_metrics["gao_biao_dt"]),
                 f"涨停{snap_metrics['zt']}跌停{snap_metrics['dt']}炸板{snap_metrics['zb']}"]
    snap_row = None
    for r in range(2, snap.max_row + 1):
        if str(snap.cell(row=r, column=1).value or "").replace("-", "") == date_str:
            snap_row = r
            break
    if snap_row:
        for i, v in enumerate(snap_vals, 1):
            snap.cell(row=snap_row, column=i, value=v)
        print(f"  行情快照 {date_str} 已就地更新")
    else:
        snap.append(snap_vals)
    _sort_sheet_by_date(snap, 14)

    # 4. 连板梯队：删除该日旧数据后重建
    lb = get_or_sheet(wb, "连板梯队")
    ensure_lb_headers(lb)
    for r in range(lb.max_row, 1, -1):
        if str(lb.cell(row=r, column=1).value or "").replace("-", "") == date_str:
            lb.delete_rows(r)
    for p in sorted(zt_pool, key=lambda x: -(int(x.get("lbc") or 0))):
        if int(p.get("lbc") or 1) < 2:
            continue
        lb.append([fmt_date(date_str), p.get("c"), p.get("n"), int(p.get("lbc") or 1),
                   p.get("hybk", ""), round(p.get("hs") or 0, 2),
                   p.get("fbt", ""), int(p.get("zbc") or 0),
                   round((p.get("ltsz") or 0) / 1e8, 1), "连板"])
    _sort_sheet_by_date(lb, 10)

    # 5. 候选验证（候选池配置中 date_str 的候选）
    cfg = load_config()
    cands = (cfg.get("候选池") or {}).get(date_str, [])
    cand_ws = get_or_sheet(wb, "候选明细")
    ensure_cand_headers(cand_ws)
    if cands:
        quotes = fetch_quotes([to_secid(c["代码"]) for c in cands])
        rows = verify_candidates(cands, zt_pool, dt_pool, quotes, fmt_date(date_str), fmt_date(prev_date))
        existing = set()
        for r in range(2, cand_ws.max_row + 1):
            if str(cand_ws.cell(row=r, column=1).value or "").replace("-", "") == date_str:
                existing.add(str(cand_ws.cell(row=r, column=4).value or ""))
        new_rows = [r for r in rows if str(r[3]) not in existing]
        for r in new_rows:
            cand_ws.append(r)
        _style_body(cand_ws, cand_ws.max_row - 1, 14, verdict_col=10)
        ok = sum(1 for r in rows if "✓" in str(r[9]))
        print(f"  候选验证: {len(rows)} 候选, 晋级 {ok} ({date_str}), 新增 {len(new_rows)}")
        # 汇总Sheet自动追加（该日已存在则跳过）
        append_graded_stats(get_or_sheet(wb, "分级命中率"), rows, date_str, prev_date, snap_metrics)
        append_direction_stats(get_or_sheet(wb, "方向命中率"), rows, date_str)
    else:
        print(f"  {date_str} 无候选配置（用 --add 添加），跳过候选验证")

    wb.save(OUT_XLSX)
    print("OK -> 数据库已刷新:", OUT_XLSX)


# ================= 计划任务 =================

def install_task():
    py = sys.executable
    script = os.path.join(BASE, "更新连板前瞻验证数据库.py")
    cmd = f'schtasks /create /tn "L8连板验证数据库" /tr "\'{py}\' \'{script}\'" /sc daily /st 15:20 /f'
    print("安装计划任务（每日 15:20 自动运行）:")
    print("  " + cmd)
    print("  Python: " + py)
    os.system(cmd)
    print("完成。如需修改时间: schtasks /change /tn L8连板验证数据库 /st HH:MM")
    print("如需删除: schtasks /delete /tn L8连板验证数据库 /f")


def show_cands():
    cfg = load_config()
    pools = cfg.get("候选池", {})
    total = sum(len(v) for v in pools.values())
    print(f"候选池配置共 {total} 条:")
    for d, lst in sorted(pools.items()):
        if lst:
            print(f"  {d}: {len(lst)} 候选")
            for c in lst:
                print(f"    {c['代码']} {c['名称']} 级别{c['级别']} 昨{c['昨板数']}板 方向[{c['方向']}] 板块[{c.get('板块','')}]")


if __name__ == "__main__":
    args = sys.argv[1:]
    if "--init-task" in args:
        install_task()
    elif "--show" in args:
        show_cands()
    elif "--add" in args:
        i = args.index("--add")
        add_candidate(args[i + 1], args[i + 2], args[i + 3], int(args[i + 4]),
                      args[i + 5], args[i + 6] if len(args) > i + 6 else datetime.date.today().strftime("%Y%m%d"),
                      args[i + 7] if len(args) > i + 7 else "")
    elif args and args[0].isdigit() and len(args[0]) == 8:
        main(args[0])
    else:
        main()